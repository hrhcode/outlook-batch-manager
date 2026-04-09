from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from outlook_batch_manager.models import (
    Account,
    AccountStatus,
    AccountSummary,
    ConnectivityStatus,
    MailCapabilityStatus,
    MailProvider,
    TokenRecord,
    TokenStatus,
)
from outlook_batch_manager.storage.database import Database


class AccountService:
    def __init__(self, database: Database) -> None:
        self.database = database

    def list_accounts(self, keyword: str = "", status: str = "") -> list[Account]:
        sql = "SELECT * FROM accounts"
        conditions: list[str] = []
        params: list[str] = []
        if keyword:
            conditions.append("(email LIKE ? OR group_name LIKE ? OR notes LIKE ?)")
            term = f"%{keyword}%"
            params.extend([term, term, term])
        if status:
            conditions.append("status = ?")
            params.append(status)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY id DESC"
        with self.database.connect() as connection:
            rows = connection.execute(sql, tuple(params)).fetchall()
        return [self._row_to_account(row) for row in rows]

    def list_account_summaries(self, keyword: str = "", status: str = "") -> list[AccountSummary]:
        sql = """
            SELECT
                accounts.*,
                tokens.status AS token_status,
                tokens.expires_at AS token_expires_at,
                tokens.refresh_token AS refresh_token
            FROM accounts
            LEFT JOIN tokens ON tokens.account_id = accounts.id
        """
        conditions: list[str] = []
        params: list[str] = []
        if keyword:
            conditions.append("(accounts.email LIKE ? OR accounts.group_name LIKE ? OR accounts.notes LIKE ?)")
            term = f"%{keyword}%"
            params.extend([term, term, term])
        if status:
            conditions.append("accounts.status = ?")
            params.append(status)
        if conditions:
            sql += " WHERE " + " AND ".join(conditions)
        sql += " ORDER BY accounts.id DESC"
        with self.database.connect() as connection:
            rows = connection.execute(sql, tuple(params)).fetchall()
        return [
            AccountSummary(
                account=self._row_to_account(row),
                token_status=row["token_status"] or "",
                token_expires_at=datetime.fromisoformat(row["token_expires_at"]) if row["token_expires_at"] else None,
                has_refresh_token=bool((row["refresh_token"] or "").strip()),
                has_client_id=bool((row["client_id_override"] or "").strip()),
            )
            for row in rows
        ]

    def get_account(self, account_id: int) -> Account | None:
        with self.database.connect() as connection:
            row = connection.execute("SELECT * FROM accounts WHERE id = ?", (account_id,)).fetchone()
        return self._row_to_account(row) if row else None

    def get_account_by_email(self, email: str) -> Account | None:
        with self.database.connect() as connection:
            row = connection.execute("SELECT * FROM accounts WHERE email = ?", (email.strip(),)).fetchone()
        return self._row_to_account(row) if row else None

    def get_token(self, account_id: int) -> TokenRecord | None:
        with self.database.connect() as connection:
            row = connection.execute("SELECT * FROM tokens WHERE account_id = ?", (account_id,)).fetchone()
        if row is None:
            return None
        return TokenRecord(
            id=int(row["id"]),
            account_id=int(row["account_id"]),
            access_token=row["access_token"],
            refresh_token=row["refresh_token"],
            expires_at=datetime.fromisoformat(row["expires_at"]) if row["expires_at"] else None,
            last_refreshed_at=datetime.fromisoformat(row["last_refreshed_at"]) if row["last_refreshed_at"] else None,
            status=TokenStatus(row["status"]),
        )

    def create_account(
        self,
        *,
        email: str,
        password: str,
        provider: str = "outlook",
        group_name: str = "default",
        notes: str = "",
        recovery_email: str = "",
        source: str = "manual",
        client_id_override: str = "",
        refresh_token: str = "",
    ) -> Account:
        account = Account(
            email=email.strip(),
            password=password.strip(),
            status=AccountStatus.PENDING,
            source=source,
            group_name=group_name.strip() or "default",
            notes=notes.strip(),
            recovery_email=recovery_email.strip(),
            mail_provider=self._normalize_provider(provider, email),
            client_id_override=client_id_override.strip(),
            import_format="manual_form",
            connectivity_status=ConnectivityStatus.CONNECTABLE if client_id_override.strip() and refresh_token.strip() else ConnectivityStatus.ACTION_REQUIRED,
            mail_capability_status=MailCapabilityStatus.NOT_READY,
        )
        saved = self.upsert_account(account)
        if refresh_token and saved.id is not None:
            self.save_token(
                TokenRecord(
                    account_id=saved.id,
                    refresh_token=refresh_token.strip(),
                    status=TokenStatus.STORED,
                )
            )
        return saved

    def upsert_account(self, account: Account) -> Account:
        now = account.created_at.isoformat(timespec="seconds") if account.created_at else self.database.now_iso()
        last_login = account.last_login_check_at.isoformat(timespec="seconds") if account.last_login_check_at else None
        last_mail_sync = account.last_mail_sync_at.isoformat(timespec="seconds") if account.last_mail_sync_at else None
        with self.database.connect() as connection:
            existing = connection.execute(
                "SELECT id, created_at FROM accounts WHERE email = ?",
                (account.email,),
            ).fetchone()
            if existing:
                connection.execute(
                    """
                    UPDATE accounts
                    SET password = ?, status = ?, source = ?, group_name = ?, notes = ?,
                        recovery_email = ?, mail_provider = ?, client_id_override = ?,
                        import_format = ?, connectivity_status = ?, mail_capability_status = ?,
                        last_login_check_at = ?, last_mail_sync_at = ?, last_error = ?
                    WHERE email = ?
                    """,
                    (
                        account.password,
                        account.status,
                        account.source,
                        account.group_name,
                        account.notes,
                        account.recovery_email,
                        account.mail_provider,
                        account.client_id_override,
                        account.import_format,
                        account.connectivity_status,
                        account.mail_capability_status,
                        last_login,
                        last_mail_sync,
                        account.last_error,
                        account.email,
                    ),
                )
                account.id = int(existing["id"])
                account.created_at = datetime.fromisoformat(existing["created_at"])
            else:
                cursor = connection.execute(
                    """
                    INSERT INTO accounts (
                        email, password, status, source, group_name, notes, recovery_email,
                        mail_provider, client_id_override, import_format, connectivity_status, mail_capability_status,
                        created_at, last_login_check_at, last_mail_sync_at, last_error
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        account.email,
                        account.password,
                        account.status,
                        account.source,
                        account.group_name,
                        account.notes,
                        account.recovery_email,
                        account.mail_provider,
                        account.client_id_override,
                        account.import_format,
                        account.connectivity_status,
                        account.mail_capability_status,
                        now,
                        last_login,
                        last_mail_sync,
                        account.last_error,
                    ),
                )
                account.id = int(cursor.lastrowid)
                account.created_at = datetime.fromisoformat(now)
        return account

    def save_token(self, token: TokenRecord) -> TokenRecord:
        expires_at = token.expires_at.isoformat(timespec="seconds") if token.expires_at else None
        refreshed = token.last_refreshed_at.isoformat(timespec="seconds") if token.last_refreshed_at else None
        with self.database.connect() as connection:
            existing = connection.execute(
                "SELECT id FROM tokens WHERE account_id = ?",
                (token.account_id,),
            ).fetchone()
            if existing:
                connection.execute(
                    """
                    UPDATE tokens
                    SET access_token = ?, refresh_token = ?, expires_at = ?, last_refreshed_at = ?, status = ?
                    WHERE account_id = ?
                    """,
                    (
                        token.access_token,
                        token.refresh_token,
                        expires_at,
                        refreshed,
                        token.status,
                        token.account_id,
                    ),
                )
                token.id = int(existing["id"])
            else:
                cursor = connection.execute(
                    """
                    INSERT INTO tokens (account_id, access_token, refresh_token, expires_at, last_refreshed_at, status)
                    VALUES (?, ?, ?, ?, ?, ?)
                    """,
                    (
                        token.account_id,
                        token.access_token,
                        token.refresh_token,
                        expires_at,
                        refreshed,
                        token.status,
                    ),
                )
                token.id = int(cursor.lastrowid)
        return token

    def update_account_runtime(
        self,
        account_id: int,
        *,
        connectivity_status: ConnectivityStatus | None = None,
        mail_capability_status: MailCapabilityStatus | None = None,
        last_login_check_at: datetime | None = None,
        last_mail_sync_at: datetime | None = None,
        last_error: str | None = None,
    ) -> Account | None:
        account = self.get_account(account_id)
        if account is None:
            return None
        if connectivity_status is not None:
            account.connectivity_status = connectivity_status
        if mail_capability_status is not None:
            account.mail_capability_status = mail_capability_status
        if last_login_check_at is not None:
            account.last_login_check_at = last_login_check_at
        if last_mail_sync_at is not None:
            account.last_mail_sync_at = last_mail_sync_at
        if last_error is not None:
            account.last_error = last_error
        return self.upsert_account(account)

    def delete_accounts(self, account_ids: list[int]) -> int:
        normalized = [int(item) for item in account_ids if item is not None]
        if not normalized:
            return 0
        placeholders = ",".join("?" for _ in normalized)
        with self.database.connect() as connection:
            connection.execute(f"DELETE FROM tokens WHERE account_id IN ({placeholders})", tuple(normalized))
            connection.execute(f"DELETE FROM messages WHERE account_id IN ({placeholders})", tuple(normalized))
            connection.execute(f"DELETE FROM mail_sync_runs WHERE account_id IN ({placeholders})", tuple(normalized))
            connection.execute(f"DELETE FROM accounts WHERE id IN ({placeholders})", tuple(normalized))
        return len(normalized)

    def export_accounts(self, destination: Path, accounts: Iterable[Account]) -> None:
        rows = [self._account_to_export_row(account) for account in accounts]
        if destination.suffix.lower() == ".csv":
            with destination.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else self._export_headers())
                writer.writeheader()
                writer.writerows(rows)
            return

        workbook = Workbook()
        worksheet = workbook.active
        headers = self._export_headers()
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row[header] for header in headers])
        workbook.save(destination)

    def import_accounts(self, source: Path, source_name: str = "import") -> int:
        if source.suffix.lower() in {".txt", ".log"}:
            return self._import_text_lines(source.read_text(encoding="utf-8").splitlines(), source_name)

        if source.suffix.lower() == ".csv":
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            if rows and self._looks_like_structured_rows(rows[0].keys()):
                return self._import_structured_rows(rows, source_name)
            handle_lines = source.read_text(encoding="utf-8-sig").splitlines()
            return self._import_text_lines(handle_lines, source_name)

        workbook = load_workbook(source, read_only=True)
        worksheet = workbook.active
        values = list(worksheet.iter_rows(values_only=True))
        if not values:
            return 0
        headers = ["" if value is None else str(value).strip() for value in values[0]]
        if self._looks_like_structured_rows(headers):
            rows = [dict(zip(headers, ["" if value is None else str(value) for value in row])) for row in values[1:]]
            return self._import_structured_rows(rows, source_name)
        raw_lines = [" ".join(str(value).strip() for value in row if value is not None).strip() for row in values]
        return self._import_text_lines(raw_lines, source_name)

    def _import_structured_rows(self, rows: list[dict[str, str]], source_name: str) -> int:
        imported = 0
        for row in rows:
            email = row.get("email", "").strip()
            password = row.get("password", "").strip()
            if not email or not password:
                continue
            refresh_token = row.get("refresh_token", "").strip()
            client_id_override = row.get("client_id_override", "").strip()
            connectivity = self._parse_connectivity_status(row.get("connectivity_status", ""))
            if connectivity == ConnectivityStatus.UNKNOWN:
                connectivity = ConnectivityStatus.CONNECTABLE if client_id_override and refresh_token else ConnectivityStatus.ACTION_REQUIRED
            account = Account(
                email=email,
                password=password,
                status=self._parse_account_status(row.get("status", "")),
                source=source_name,
                group_name=row.get("group_name", "default").strip() or "default",
                notes=row.get("notes", "").strip(),
                recovery_email=row.get("recovery_email", "").strip(),
                mail_provider=self._normalize_provider(row.get("mail_provider", ""), email),
                client_id_override=client_id_override,
                import_format=row.get("import_format", "table").strip() or "table",
                connectivity_status=connectivity,
                mail_capability_status=MailCapabilityStatus.NOT_READY,
            )
            saved = self.upsert_account(account)
            if refresh_token and saved.id is not None:
                self.save_token(
                    TokenRecord(
                        account_id=saved.id,
                        refresh_token=refresh_token,
                        status=TokenStatus.STORED,
                    )
                )
            imported += 1
        return imported

    def _import_text_lines(self, lines: list[str], source_name: str) -> int:
        imported = 0
        for line in lines:
            parsed = self._parse_text_account(line, source_name)
            if parsed is None:
                continue
            account, refresh_token = parsed
            saved = self.upsert_account(account)
            if refresh_token and saved.id is not None:
                self.save_token(
                    TokenRecord(
                        account_id=saved.id,
                        refresh_token=refresh_token,
                        status=TokenStatus.STORED,
                    )
                )
            imported += 1
        return imported

    @staticmethod
    def _looks_like_structured_rows(headers: Iterable[str]) -> bool:
        normalized = {header.strip().lower() for header in headers if header}
        return {"email", "password"}.issubset(normalized)

    def _parse_text_account(self, line: str, source_name: str) -> tuple[Account, str] | None:
        clean = line.strip()
        if not clean or clean.lower().startswith("email"):
            return None
        parts = [part.strip() for part in re.split(r"-{4,}", clean) if part.strip()]
        if len(parts) < 2:
            return None
        email, password = parts[0], parts[1]
        client_id_override = parts[2] if len(parts) >= 3 else ""
        refresh_token = parts[3] if len(parts) >= 4 else ""
        account = Account(
            email=email,
            password=password,
            status=AccountStatus.PENDING,
            source=source_name,
            group_name="default",
            notes="",
            recovery_email="",
            mail_provider=self._normalize_provider("", email),
            client_id_override=client_id_override,
            import_format="token_line" if refresh_token else "pair_line",
            connectivity_status=ConnectivityStatus.CONNECTABLE if client_id_override and refresh_token else ConnectivityStatus.ACTION_REQUIRED,
            mail_capability_status=MailCapabilityStatus.NOT_READY,
        )
        return account, refresh_token

    @staticmethod
    def _export_headers() -> list[str]:
        return [
            "email",
            "password",
            "status",
            "source",
            "group_name",
            "notes",
            "recovery_email",
            "mail_provider",
            "client_id_override",
            "import_format",
            "connectivity_status",
            "mail_capability_status",
            "created_at",
            "last_login_check_at",
            "last_mail_sync_at",
            "last_error",
            "refresh_token",
        ]

    def _account_to_export_row(self, account: Account) -> dict[str, str]:
        token = self.get_token(account.id or 0) if account.id else None
        return {
            "email": account.email,
            "password": account.password,
            "status": account.status,
            "source": account.source,
            "group_name": account.group_name,
            "notes": account.notes,
            "recovery_email": account.recovery_email,
            "mail_provider": account.mail_provider,
            "client_id_override": account.client_id_override,
            "import_format": account.import_format,
            "connectivity_status": account.connectivity_status,
            "mail_capability_status": account.mail_capability_status,
            "created_at": account.created_at.isoformat(timespec="seconds") if account.created_at else "",
            "last_login_check_at": account.last_login_check_at.isoformat(timespec="seconds") if account.last_login_check_at else "",
            "last_mail_sync_at": account.last_mail_sync_at.isoformat(timespec="seconds") if account.last_mail_sync_at else "",
            "last_error": account.last_error,
            "refresh_token": token.refresh_token if token else "",
        }

    def _row_to_account(self, row) -> Account:
        return Account(
            id=int(row["id"]),
            email=row["email"],
            password=row["password"],
            status=AccountStatus(row["status"]),
            source=row["source"],
            group_name=row["group_name"],
            notes=row["notes"],
            recovery_email=row["recovery_email"],
            mail_provider=MailProvider(row["mail_provider"] or "outlook"),
            client_id_override=row["client_id_override"] or "",
            import_format=row["import_format"] or "manual",
            connectivity_status=ConnectivityStatus(row["connectivity_status"] or ConnectivityStatus.UNKNOWN),
            mail_capability_status=MailCapabilityStatus(row["mail_capability_status"] or MailCapabilityStatus.NOT_READY),
            created_at=datetime.fromisoformat(row["created_at"]),
            last_login_check_at=datetime.fromisoformat(row["last_login_check_at"]) if row["last_login_check_at"] else None,
            last_mail_sync_at=datetime.fromisoformat(row["last_mail_sync_at"]) if row["last_mail_sync_at"] else None,
            last_error=row["last_error"] or "",
        )

    def _normalize_provider(self, provider: str, email: str) -> MailProvider:
        lower = provider.strip().lower()
        if lower in {MailProvider.OUTLOOK, MailProvider.HOTMAIL}:
            return MailProvider(lower)
        email_lower = email.lower()
        if "@hotmail." in email_lower:
            return MailProvider.HOTMAIL
        return MailProvider.OUTLOOK

    def _parse_account_status(self, value: str) -> AccountStatus:
        normalized = value.strip().lower()
        if not normalized:
            return AccountStatus.PENDING
        return AccountStatus(normalized)

    def _parse_connectivity_status(self, value: str) -> ConnectivityStatus:
        normalized = value.strip().lower()
        if not normalized:
            return ConnectivityStatus.UNKNOWN
        return ConnectivityStatus(normalized)
